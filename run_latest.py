from ntc_templates.parse import parse_output
from netmiko import ConnectHandler
import json
from openpyxl import *
import xlwt
import io
import textfsm



# -----------Load Workbook/Worksheet-----------
workbook = load_workbook("helper.xlsx")
worksheet = workbook["python"]
worksheet1=workbook.create_sheet("svi")
worksheet1.cell(1,1).value="VLAN-ID"
worksheet1.cell(1,2).value="VLAN-NAME"
worksheet1.cell(1,3).value="SVI-ADDRESS"
worksheet1.cell(1,4).value="SUBNET-MASK"
worksheet1.cell(1,5).value="COMMENT"
# -----------Load Workbook/Worksheet-----------

# -----------Connect Device-----------
deviceConnection = {
    'device_type': 'cisco_ios',
    'host': '192.168.138.10',
    'username': 'keremcan',
    'password': 'onemli',
    'port': 22,
    'secret': ''}
connect = ConnectHandler(**deviceConnection)
# -----------Connect Device-----------


# -----------Interface Brief----------
interface_brief_output = connect.send_command("show ip interface brief")
interface_parsed = parse_output(platform="cisco_ios", command="show ip interface brief", data=interface_brief_output)
# -----------Interface Brief----------




# -----------L2/L3 Info Function----------
# def switchport(interface):
#     interfaces_switchport_output = connect.send_command("""show interfaces """ + interface + """ switchport""")
#     interfaces_switchport_parsed = parse_output(platform="cisco_ios", command="show interfaces switchport",
#                                                 data=interfaces_switchport_output)
#     for item in interfaces_switchport_parsed:
#         key = get_key(item["interface"],interfaces_dictionary)
#         worksheet.cell(key, 3).value=item["switchport"]
#
#     workbook.save("helper.xlsx")
# -----------L2/L3 Info Function----------

# -----------Interface Value to Key Function ----------
def get_key(val, dictionary):
    for key, value in dictionary.items():
        if value == val:
            return key
            # return key değeri excel'de interface'in bulunduğu row değerini döner
    return "Error"
# -----------Interface Value to Key Function ----------



row = 2
interfaces_dictionary = {}
short_interfaces_dictionary = {}
for item in interface_parsed:

    if item["intf"].startswith("FastEthernet"):
        worksheet.cell(row, 1).value = item["intf"].replace("FastEthernet", "Fa")
        worksheet.cell(row, 2).value = item["status"]
        interfaces_dictionary[row] = item["intf"]
        short_interfaces_dictionary[row] = item["intf"].replace("FastEthernet", "Fa")
        row += 1
    elif item["intf"].startswith("GigabitEthernet"):
        worksheet.cell(row, 1).value = item["intf"].replace("GigabitEthernet","Gi")
        worksheet.cell(row, 2).value = item["status"]
        interfaces_dictionary[row] = item["intf"]
        short_interfaces_dictionary[row] = item["intf"].replace("GigabitEthernet","Gi")
        row += 1
    elif item["intf"].startswith("TwoGigabitEthernet"):
        worksheet.cell(row, 1).value = item["intf"].replace("TwoGigabitEthernet", "Tw")
        worksheet.cell(row, 2).value = item["status"]
        interfaces_dictionary[row] = item["intf"]
        short_interfaces_dictionary[row] = item["intf"].replace("TwoGigabitEthernet", "Tw")
        row += 1
    elif item["intf"].startswith("FiveGigabitEthernet"):
        worksheet.cell(row, 1).value = item["intf"].replace("FiveGigabitEthernet", "Fi")
        worksheet.cell(row, 2).value = item["status"]
        interfaces_dictionary[row] = item["intf"]
        short_interfaces_dictionary[row] = item["intf"].replace("FiveGigabitEthernet", "Fi")
        row += 1
    elif item["intf"].startswith("TenGigabitEthernet"):
        worksheet.cell(row, 1).value = item["intf"].replace("TenGigabitEthernet", "Te")
        worksheet.cell(row, 2).value = item["status"]
        interfaces_dictionary[row] = item["intf"]
        short_interfaces_dictionary[row] = item["intf"].replace("TenGigabitEthernet", "Te")
        row += 1
    elif item["intf"].startswith("TwentyFiveGig"):
        worksheet.cell(row, 1).value = item["intf"].replace("TwentyFiveGig", "Twe")
        worksheet.cell(row, 2).value = item["status"]
        interfaces_dictionary[row] = item["intf"]
        short_interfaces_dictionary[row] = item["intf"].replace("TwentyFiveGig", "Twe")
        row += 1
    elif item["intf"].startswith("Ethernet"):
        worksheet.cell(row, 1).value = item["intf"].replace("Ethernet", "Eth")
        worksheet.cell(row, 2).value = item["status"]
        interfaces_dictionary[row] = item["intf"]
        short_interfaces_dictionary[row] = item["intf"].replace("Ethernet", "Eth")
        row += 1

    elif item["intf"].startswith("Port-channel"):
        worksheet.cell(row, 1).value = item["intf"].replace("Port-channel", "Po")
        worksheet.cell(row, 2).value = item["status"]
        interfaces_dictionary[row] = item["intf"]
        short_interfaces_dictionary[row] = item["intf"].replace("Port-channel", "Po")
        row += 1
    elif item["intf"].startswith("Vlan"):
        worksheet.cell(row, 1).value = item["intf"].replace("Vlan", "Vl")
        worksheet.cell(row, 2).value = item["status"]
        interfaces_dictionary[row] = item["intf"]
        short_interfaces_dictionary[row] = item["intf"].replace("Vlan", "Vl")
        row += 1
workbook.save("helper.xlsx")
print(interfaces_dictionary)
print(interfaces_dictionary[2])



row=2
for int_col_num in range(2, len(interfaces_dictionary) + 1):
    interface=interfaces_dictionary[int_col_num]
    try:
        sendcommand="""show interfaces """+interface+""" switchport"""
        interfaces_switchport_output = connect.send_command(str(sendcommand))
        command_str="""show interfaces switchport"""

        interfaces_switchport_parsed = parse_output(platform="cisco_ios", command=str(command_str), data=interfaces_switchport_output)
    except:
        print("Error Detected")

    for item in interfaces_switchport_parsed:
        if item["switchport"]=="Enabled":
            worksheet.cell(row, 3).value = "L2"
        elif item["switchport"]=="Disabled":
            worksheet.cell(row, 3).value = "L3"
    row+=1
workbook.save("helper.xlsx")

# -----------Cdp Neighbors Detail------
cdp_neigh_det_output = connect.send_command("show cdp neighbors detail")
cdp_neigh_det_parsed = parse_output(platform="cisco_ios", command="show cdp neighbors detail", data= cdp_neigh_det_output)

print(json.dumps(cdp_neigh_det_parsed,indent=4))

for item in cdp_neigh_det_parsed:
    cdp=get_key(item["local_port"],interfaces_dictionary)
    if cdp!="Error":
        worksheet.cell(cdp, 4).value = item["destination_host"]
        worksheet.cell(cdp, 5).value = item["remote_port"]
    cdp=""
workbook.save("helper.xlsx")
# -----------Cdp Neighbors Detail------

# -----------Lldp Neighbors Detail Reel Cihazlarla Denenecek------
# cdp_neigh_det_output = connect.send_command("show lldp neighbors detail")
# cdp_neigh_det_parsed = parse_output(platform="cisco_ios", command="show lldp neighbors detail", data= cdp_neigh_det_output)
#
# print(json.dumps(cdp_neigh_det_parsed,indent=4))
#
# for item in cdp_neigh_det_parsed:
#     cdp=get_key(item["local_port"],interfaces_dictionary)
#     if cdp!="Error":
#         worksheet.cell(cdp, 4).value = item["destination_host"]
#         worksheet.cell(cdp, 5).value = item["remote_port"]
#     cdp=""
# workbook.save("helper.xlsx")
# -----------Lldp Neighbors Detail Reel Cihazlarla Denenecek------


# -----------Etherchannel Summary------
eth_summary_output = connect.send_command("show etherchannel summary")
eth_summary_parsed = parse_output(platform="cisco_ios", command="show etherchannel summary", data= eth_summary_output)
print(json.dumps(eth_summary_parsed,indent=4)) # burası print kısmı sonra kaldırılacak

for item in eth_summary_parsed:
    for interface_item in item["interfaces"]:
        po=get_key(interface_item, short_interfaces_dictionary)
        worksheet.cell(po, 6).value = item["po_name"]
        sh_run_int = connect.send_command("""show run interface"""+interface_item)
        if sh_run_int.find("mode active"):
            worksheet.cell(po, 7).value = "Active"
        elif sh_run_int.find("mode one"):
            worksheet.cell(po, 7).value = "On"

workbook.save("helper.xlsx")
# -----------Etherchannel Summary------






# -----------Allowed Vlan------

#------------Allowed Vlan Parser TextFSM
allowed_vlan_parser_template = io.StringIO("""\
Value Port (\S+(/\d+)?)
Value Vlans (\d+([-,]\d+)+)

Start
  ^Port\s+Vlans allowed on trunk$$ -> Begin

Begin
  ^${Port}\s+${Vlans}$$ -> Record
  ^Port\s+Vlans allowed and active in management domain$$ -> End
""")


#------------Allowed Vlan Parser TextFSM

for int_col_num in range(2, len(short_interfaces_dictionary) + 2):
    interface = short_interfaces_dictionary[int_col_num]

    print(interface)
    allowed_vlan_output = connect.send_command("""show interface """ + interface + """ trunk""")

    fsm = textfsm.TextFSM(allowed_vlan_parser_template)
    result = fsm.ParseText(allowed_vlan_output)
    row_id=get_key(interface,short_interfaces_dictionary)
    try:
        print(result[0][1])
        worksheet.cell(row_id, 8).value = result[0][1]
        worksheet.cell(row_id, 9).value = "Trunk"

    except:
        worksheet.cell(row_id, 9).value = "Access"
        print("Have no result")
        worksheet.cell(row_id, 8).value = "No Result"



workbook.save("helper.xlsx")
# -----------Allowed Vlan------

#------------Interfaces Desc----
int_desc_output = connect.send_command("""show interfaces description""")
int_desc_parsed = parse_output(platform="cisco_ios", command="show interfaces description", data= int_desc_output)
print(json.dumps(int_desc_parsed,indent=4))
for item in int_desc_parsed:
    row_id=get_key(item["port"],short_interfaces_dictionary)
    try:
        worksheet.cell(row_id, 10).value = item["descrip"]
    except:
        worksheet.cell(row_id, 10).value = "No Description"
workbook.save("helper.xlsx")
#------------Interfaces Desc----

#------------SVI----

ip_int_output =connect.send_command("""show ip interface brief""")
ip_int_parsed =parse_output(platform="cisco_ios", command="show ip interface brief", data= ip_int_output)
#print(json.dumps(ip_int_parsed ,indent=4))
sh_vlan_output = connect.send_command("""show vlan brief""")
sh_vlan_parsed  = parse_output(platform="cisco_ios", command="show vlan brief", data= sh_vlan_output)
row=2
for item in sh_vlan_parsed :

    worksheet1.cell(row,1).value=item["vlan_id"]
    worksheet1.cell(row,2).value=item["name"]
    for interface in ip_int_parsed:
        x=interface["intf"]
        try:
            if int(item["vlan_id"])==int(x[4:]):
                worksheet1.cell(row, 3).value=interface["ipaddr"]
                break
        except:
            print("Error - - - - This is not SVI")
    row += 1
workbook.save("helper.xlsx")
#print(json.dumps(sh_vlan_parsed ,indent=4))

#------------SVI----
