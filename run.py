from ntc_templates.parse import parse_output
from netmiko import ConnectHandler
import json
from openpyxl import *
import xlwt
# -----------Load Workbook/Worksheet-----------
workbook = load_workbook("helper.xlsx")
worksheet = workbook["python"]
# -----------Load Workbook/Worksheet-----------

# -----------Connect Device-----------
deviceConnection = {
    'device_type': 'cisco_ios',
    'host': '192.168.138.10',
    'username': 'keremcan',
    'password': 'onemli',
    'port': 22,
    'secret': ''}
connect = ConnectHandler(**deviceConnection)
# -----------Connect Device-----------


# -----------Interface Brief----------
interface_brief_output = connect.send_command("show ip interface brief")
interface_parsed = parse_output(platform="cisco_ios", command="show ip interface brief", data=interface_brief_output)
# -----------Interface Brief----------




# -----------L2/L3 Info Function----------
# def switchport(interface):
#     interfaces_switchport_output = connect.send_command("""show interfaces """ + interface + """ switchport""")
#     interfaces_switchport_parsed = parse_output(platform="cisco_ios", command="show interfaces switchport",
#                                                 data=interfaces_switchport_output)
#     for item in interfaces_switchport_parsed:
#         key = get_key(item["interface"],interfaces_dictionary)
#         worksheet.cell(key, 3).value=item["switchport"]
#
#     workbook.save("helper.xlsx")
# -----------L2/L3 Info Function----------

# -----------Interface Value to Key Function ----------
#def get_key(val, dictionary):
#    for key, value in dictionary.items():
#        if value == val:
#            return key
#            # return key değeri excel'de interface'in bulunduğu row değerini döner
#    return "key doesn't exist"
# -----------Interface Value to Key Function ----------



row = 2
interfaces_dictionary = {}
for item in interface_parsed:
    worksheet.cell(row, 1).value = item["intf"]
    worksheet.cell(row, 2).value = item["status"]
    interfaces_dictionary[row] = item["intf"]
    row += 1
workbook.save("helper.xlsx")
print(interfaces_dictionary)
print(interfaces_dictionary[2])



row=2
for int_col_num in range(2, len(interfaces_dictionary) + 1):
    interface=interfaces_dictionary[int_col_num]
    sendcommand="""show interfaces """+interface+""" switchport"""
    interfaces_switchport_output = connect.send_command(str(sendcommand))
    command_str="""show interfaces switchport"""
    interfaces_switchport_parsed = parse_output(platform="cisco_ios", command=str(command_str), data=interfaces_switchport_output)

    for item in interfaces_switchport_parsed:
        if item["switchport"]=="Enabled":
            worksheet.cell(row, 3).value = "L2"
        elif item["switchport"]=="Disabled":
            worksheet.cell(row, 3).value = "L3"
    row+=1
workbook.save("helper.xlsx")

# -----------Cdp Neighbors Detail------
cdp_neigh_det_output = connect.send_command("show cdp neighbors detail")
cdp_neigh_det_parsed = parse_output(platform="cisco_ios", command="show cdp neighbors detail", data= cdp_neigh_det_output)

print(json.dumps(cdp_neigh_det_parsed,indent=4))



# -----------Cdp Neighbors Detail------
